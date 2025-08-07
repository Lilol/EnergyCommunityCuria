import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows

# File paths
pod_file_path = 'C:\\NextCloud\\Curia\\files\\input\\DatiComuni\\SettimoTorinese\\lista_pod.csv'
bill_file_path = 'C:\\NextCloud\\Curia\\files\\input\\DatiComuni\\SettimoTorinese\\dati_bollette.csv'
output_pod_check_path = 'C:\\NextCloud\\Curia\\files\\input\\DatiComuni\\SettimoTorinese\\lista_pod_check.xlsx'
output_bill_check_path = 'C:\\NextCloud\\Curia\\files\\input\\DatiComuni\\SettimoTorinese\\dati_bollette_check.xlsx'

# Load CSVs
pod_df = pd.read_csv(pod_file_path, sep=';')
bill_df = pd.read_csv(bill_file_path, sep=';')

# Normalize column names
pod_df.columns = [col.strip().lower() for col in pod_df.columns]
bill_df.columns = [col.strip().lower() for col in bill_df.columns]

# Fill colors
red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")       # Critical errors
orange_fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")    # Warnings
yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")    # Attention

# ----- POD FILE CHECK -----
wb_pod = Workbook()
ws_pod = wb_pod.active
ws_pod.title = "POD Check"

# Write DataFrame to Excel
for r in dataframe_to_rows(pod_df, index=False, header=True):
    ws_pod.append(r)

# Highlight duplicate PODs
if 'pod' in pod_df.columns:
    duplicate_pods = pod_df['pod'].duplicated(keep=False)
    for row_idx, is_dup in enumerate(duplicate_pods, start=2):
        if is_dup:
            ws_pod.cell(row=row_idx, column=pod_df.columns.get_loc('pod') + 1).fill = red_fill

# ----- BILL FILE CHECK -----
wb_bill = Workbook()
ws_bill = wb_bill.active
ws_bill.title = "Bill Check"

# Write DataFrame to Excel
for r in dataframe_to_rows(bill_df, index=False, header=True):
    ws_bill.append(r)

# Valid PODs from pod list
valid_pods = set(pod_df['pod'].dropna().unique()) if 'pod' in pod_df.columns else set()

# Check conditions
for idx, row in bill_df.iterrows():
    excel_row = idx + 2  # Excel is 1-indexed, header is row 1

    if pd.isna(bill_df.loc[idx, 'f0']):
        for f in ['f1', 'f2', 'f3']:
            if pd.isna(bill_df.loc[idx, f]):
                ws_bill.cell(row=excel_row + 2, column=bill_df.columns.get_loc('f0') + 1).fill = red_fill
                ws_bill.cell(row=excel_row + 2, column=bill_df.columns.get_loc(f) + 1).fill = red_fill

    for col_idx, col in enumerate(bill_df.columns, start=1):
        cell_value = row[col]

        # Check 1: Empty or NaN
        if col not in ("f0", "f1", "f2", "f3") and pd.isna(cell_value) or (isinstance(cell_value, str) and cell_value.strip() == ''):
            ws_bill.cell(row=excel_row, column=col_idx).fill = red_fill

        # Check 2: Invalid POD
        if col == 'pod':
            if cell_value not in valid_pods:
                ws_bill.cell(row=excel_row, column=col_idx).fill = red_fill

        # Check 3: Duplicate POD
        if col == 'pod' and bill_df[['pod', 'mese']].duplicated(keep=False)[idx]:
            ws_bill.cell(row=excel_row, column=col_idx).fill = orange_fill

        # Check 4: Negative values (e.g. amount fields)
        if pd.api.types.is_numeric_dtype(bill_df[col]) and not pd.isna(cell_value) and cell_value < 0:
            ws_bill.cell(row=excel_row, column=col_idx).fill = yellow_fill

# Save both workbooks
wb_pod.save(output_pod_check_path)
wb_bill.save(output_bill_check_path)

