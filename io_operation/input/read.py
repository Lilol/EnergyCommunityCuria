import logging
import os
from os.path import join, exists

import numpy as np
import xarray as xr
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from pandas import read_csv, read_excel, to_datetime, DataFrame, isna, api, notna

from data_processing_pipeline.definitions import Stage
from data_storage.data_store import DataStore
from data_storage.omnes_data_array import OmnesDataArray
from io_operation.input.definitions import DataKind, PvDataSource
from io_operation.input.file_checker import red_fill, yellow_fill, orange_fill
from io_operation.io_operation_separately_by_attribute import IoOperationSeparately
from utility import configuration
from utility.definitions import append_extension
from utility.enum_definitions import convert_value_to_enum

logger = logging.getLogger(__name__)


class Read(IoOperationSeparately):
    stage = Stage.READ
    _column_names = convert_value_to_enum
    _name = "reader"
    _input_root = configuration.config.get("path", "input")
    _directory = ""
    _filename = ""
    _ext = ".csv"

    def __init__(self, name=_name, *args, **kwargs):
        super().__init__(name, *args, **kwargs)
        self._filename = kwargs.pop("filename", self.__class__._filename)
        self._directory = kwargs.pop("directory", self.__class__._directory)
        self._input_root = kwargs.pop("input_root", self.__class__._input_root)
        self._path = join(self._input_root, self._directory)
        self._data = None

    def _io_operation(self, dataset: OmnesDataArray | None, attribute=None, attribute_value="", *args,
                      **kwargs) -> OmnesDataArray | None:
        filename = os.path.join(self._path, attribute_value, append_extension(self._filename, self._ext))
        if not exists(filename):
            logger.warning(f"File {filename} does not exist, skipping file reading.")
            return dataset
        data = self.read_data(filename, attribute, attribute_value)
        if self._data is None:
            self._data = data
        else:
            self._data = xr.concat([self._data, data], dim=attribute)
        return self._data

    def execute(self, dataset: OmnesDataArray | None, *args, **kwargs) -> OmnesDataArray | None:
        municipalities = kwargs.pop("municipality", configuration.config.get("rec", "municipalities"))
        return super().execute(dataset, separate_to_directories_by=DataKind.MUNICIPALITY.value,
                               directories=municipalities, *args, **kwargs)

    def read_data(self, filename, attribute, attribute_value):
        data = read_csv(filename, sep=';', parse_dates=True).rename(columns=self.__class__._column_names)
        self.validate(filename, data, attribute, attribute_value)
        return OmnesDataArray(data=data.reset_index(drop=True)).expand_dims(attribute).assign_coords(
            {attribute: [attribute_value]})

    def validate(self, filename, data, attribute, attribute_value):
        pass


class ReadDataArray(Read):
    _input_root = configuration.config.get("path", "output")
    _name = "data_array_reader"
    _ext = ".nc"

    def read_data(self, filename, attribute, attribute_value):
        if not exists(filename):
            return OmnesDataArray()
        dataset = xr.open_dataarray(filename, engine="netcdf4")
        dataset = dataset.assign_coords(
            {dim: [convert_value_to_enum(coord) for coord in dataset[dim].values] for dim in dataset.dims})
        if DataKind.TIME.value in dataset.coords:
            dataset = dataset.assign_coords(time=to_datetime(dataset.time.values))
        return dataset


class ReadProduction(Read):
    _name = "production_reader"

    def __init__(self, name=_name, *args, **kwargs):
        super().__init__(name, *args, **kwargs)
        self._pv_resource_reader = ReadPvProduction.create()

    def execute(self, dataset: OmnesDataArray | None, *args, **kwargs) -> OmnesDataArray | None:
        municipalities = kwargs.pop("municipality", configuration.config.get("rec", "municipalities"))
        user_data = DataStore()["pv_plants"]
        self._data = xr.concat(
            [self._pv_resource_reader.execute(self._data, municipality=municipality, user=user) for municipality in
             municipalities for user in
             user_data.sel({DataKind.MUNICIPALITY.value: municipality})[DataKind.USER.value].values],
            dim=DataKind.USER.value)
        return self._data


class ReadPvProduction(Read):
    _name = "pvresource_reader"
    _directory = "DatiComuni"
    _ref_year = configuration.config.getint("time", "year")

    @classmethod
    def create(cls, *args, **kwargs):
        source = configuration.config.get("production", "estimator")
        return ReadPvgis(*args, **kwargs) if source == PvDataSource.PVGIS else ReadPvSol(*args, **kwargs)

    @classmethod
    def replace_year(cls, df):
        df.index = df.index.map(lambda x: x.replace(year=cls._ref_year))
        return df


class ReadPvgis(ReadPvProduction):
    _name = "pvgis_reader"
    _production_column_name = 'Irradiance onto horizontal plane '  # hourly production of the plants (kW)
    _column_names = {"power": DataKind.POWER}

    def execute(self, dataset: OmnesDataArray | None, *args, **kwargs) -> OmnesDataArray | None:
        municipality = kwargs.pop("municipality", configuration.config.get("rec", "location"))
        user = kwargs.pop("user")
        production = read_csv(join(self._path, municipality, PvDataSource.PVGIS.value, f"{user}.csv"), sep=';',
                              index_col=0, parse_dates=True, date_format="%d/%m/%Y %H:%M").rename(
            columns=self._column_names)
        production = self.replace_year(production)
        production = OmnesDataArray(data=production).rename(
            {"dim_1": DataKind.POWER.value, "timestamp": DataKind.TIME.value}).expand_dims(
            [DataKind.MUNICIPALITY.value, DataKind.USER.value]).assign_coords(
            {DataKind.MUNICIPALITY.value: [municipality, ], DataKind.USER.value: [user, ]}).squeeze(
            DataKind.POWER.value, drop=True)
        return production


class ReadPvSol(ReadPvProduction):
    _name = "pvsol_reader"
    _production_column_name = 'Grid Export '  # hourly production of the plants (kW)
    _column_names = {_production_column_name: DataKind.POWER}

    def execute(self, dataset: OmnesDataArray | None, *args, **kwargs) -> OmnesDataArray | None:
        municipality = kwargs.pop("municipality", configuration.config.get("rec", "location"))
        user = kwargs.pop("user", "")
        production = read_csv(join(self._path, municipality, PvDataSource.PVSOL.value, f"{user}.csv"), sep=';',
                              decimal=',', skiprows=range(1, 17), index_col=0, header=0, parse_dates=True,
                              date_format="%d.%m. %H:%M", usecols=["Time", self._production_column_name]).rename(
            columns=self._column_names)
        production = self.replace_year(production)
        production = OmnesDataArray(data=production).rename(
            {"dim_1": DataKind.POWER.value, "Time": DataKind.TIME.value}).expand_dims(
            [DataKind.MUNICIPALITY.value, DataKind.USER.value]).assign_coords(
            {DataKind.MUNICIPALITY.value: [municipality, ], DataKind.USER.value: [user, ]}).squeeze(
            DataKind.POWER.value, drop=True)
        return production


class ReadPvPlantData(Read):
    _name = "pv_plant_reader"

    _column_names = {'pod': DataKind.USER,  # code or name of the associated end user
                     'descrizione': DataKind.DESCRIPTION,  # description of the associated end user
                     'indirizzo': DataKind.USER_ADDRESS,  # address of the end user
                     'pv_size': DataKind.POWER,  # size of the plant (kW)
                     'produzione annua [kWh]': DataKind.ANNUAL_ENERGY,  # annual energy produced (kWh)
                     'rendita specifica [kWh/kWp]': DataKind.ANNUAL_YIELD,  # specific annual production (kWh/kWp)
                     }

    _directory = "DatiComuni"
    _filename = "lista_impianti.csv"

    def __init__(self, name=_name, *args, **kwargs):
        super().__init__(name, *args, **kwargs)


class ReadUserData(Read):
    _name = "users_reader"
    _column_names = {'pod': DataKind.USER,  # code or name of the end user
                     'descrizione': DataKind.DESCRIPTION,  # description of the end user
                     'indirizzo': DataKind.USER_ADDRESS,  # address of the end user
                     'tipo': DataKind.USER_TYPE,  # type of end user
                     'potenza': DataKind.POWER,  # maximum available power of the end-user (kW)
                     }

    _directory = "DatiComuni"
    _filename = "lista_pod.csv"  # list of end-users

    def __init__(self, name=_name, *args, **kwargs):
        super().__init__(name, *args, **kwargs)

    def _validate_and_export_excel(self, df: DataFrame, filename: str):
        wb = Workbook()
        ws = wb.active
        ws.title = "POD Check"

        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)

        if 'pod' in df.columns:
            duplicate_pods = df['pod'].duplicated(keep=False)
            for row_idx, is_dup in enumerate(duplicate_pods, start=2):
                if is_dup:
                    ws.cell(row=row_idx, column=df.columns.get_loc('pod') + 1).fill = red_fill

        output_path = os.path.splitext(filename)[0] + "_check.xlsx"
        wb.save(output_path)


class ReadBills(Read):
    _name = "bill_reader"
    _time_of_use_energy_column_names = {f'f{i}': tou_energy_name for i, tou_energy_name in
                                        enumerate(configuration.config.get("tariff", "time_of_use_labels"), 1)}

    _column_names = {'pod': DataKind.USER,  # code or name of the end user
                     'anno': DataKind.YEAR,  # year
                     'mese': DataKind.MONTH,  # number of the month
                     'totale': DataKind.ANNUAL_ENERGY,  # Annual consumption
                     'f0': DataKind.MONO_TARIFF, **_time_of_use_energy_column_names}

    _directory = "DatiComuni"
    _filename = "dati_bollette.csv"

    def __init__(self, name=_name, *args, **kwargs):
        super().__init__(name, *args, **kwargs)

    def execute(self, dataset: OmnesDataArray | None, *args, **kwargs) -> OmnesDataArray | None:
        super().execute(dataset, *args, **kwargs)
        # Check that each user has exactly 12 rows in the bills dataframe
        users = self._data.sel({"dim_1": DataKind.USER})
        if not (np.all(users.groupby(users).count() == 12)).all():
            logger.warning(
                "All end users in 'data_users_bills' must have exactly 12 rows, but a user is found with more or less"
                " rows.")
        # Time of use labels
        configuration.config.set_and_check("tariff", "time_of_use_labels", self._data["dim_1"][
            self._data["dim_1"].isin(list(self._time_of_use_energy_column_names.values()))].values,
                                           setter=configuration.config.setarray, check=False)
        return self._data

    def _validate_and_export_excel(self, df: DataFrame, filename: str):
        # Load valid PODs for check (e.g., from lista_pod.csv or internal cache)
        pod_path = os.path.join(self._input_root, "DatiComuni", "lista_pod.csv")
        if os.path.exists(pod_path):
            pod_df = read_csv(pod_path, sep=';')
            pod_df.columns = [col.strip().lower() for col in pod_df.columns]
            valid_pods = set(pod_df['pod'].dropna().unique())
        else:
            valid_pods = set()

        wb = Workbook()
        ws = wb.active
        ws.title = "Bill Check"

        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)

        for idx, row in df.iterrows():
            excel_row = idx + 2

            if isna(row.get('f0')):
                for f in ['f1', 'f2', 'f3']:
                    if isna(row.get(f)):
                        ws.cell(row=excel_row, column=df.columns.get_loc('f0') + 1).fill = red_fill
                        ws.cell(row=excel_row, column=df.columns.get_loc(f) + 1).fill = red_fill

            for col_idx, col in enumerate(df.columns, start=1):
                val = row[col]
                if col not in ("f0", "f1", "f2", "f3") and (isna(val) or (isinstance(val, str) and val.strip() == '')):
                    ws.cell(row=excel_row, column=col_idx).fill = red_fill
                if col == 'pod':
                    if val not in valid_pods:
                        ws.cell(row=excel_row, column=col_idx).fill = red_fill
                    if df[['pod', 'mese']].duplicated(keep=False)[idx]:
                        ws.cell(row=excel_row, column=col_idx).fill = orange_fill
                if api.types.is_numeric_dtype(df[col]) and notna(val) and val < 0:
                    ws.cell(row=excel_row, column=col_idx).fill = yellow_fill

        output_path = os.path.splitext(filename)[0] + "_check.xlsx"
        wb.save(output_path)


class ReadCommonData(Read):
    _column_names = {}

    def execute(self, dataset: OmnesDataArray | None, *args, **kwargs) -> OmnesDataArray | None:
        self._data = read_csv(join(self._path, self._filename), sep=';', index_col=0, header=0).rename(
            columns=self._column_names)
        return OmnesDataArray(data=self._data)


class ReadTariff(ReadCommonData):
    # ARERA's division depending on subdivision into tariff time-slots
    # NOTE : f : 1 - tariff time-slot F1, central hours of work-days
    #            2 - tariff time-slot F2, evening of work-days, and saturdays
    #            3 - tariff times-lot F2, night, sundays and holidays
    _name = "tariff_reader"
    _directory = "Common"
    _filename = "arera.csv"

    def __init__(self, name=_name, *args, **kwargs):
        super().__init__(name, *args, **kwargs)


class ReadTypicalLoadProfile(ReadCommonData):
    _name = "typical_load_profile_reader"
    _column_names = {'type': DataKind.USER_TYPE,  # code or name of the end user
                     'month': DataKind.MONTH}
    _directory = "Common"
    _filename = "y_ref_gse.csv"

    # Reference profiles from GSE
    def __init__(self, name=_name, *args, **kwargs):
        super().__init__(name, *args, **kwargs)


class ReadGseDatabase(ReadTypicalLoadProfile):
    _name = "gse_database_reader"
    _column_names = convert_value_to_enum
    _directory = "DatabaseGSE"
    _filename = "gse_ref_profiles.xlsx"

    # Reference profiles from GSE
    def __init__(self, name=_name, *args, **kwargs):
        super().__init__(name, *args, **kwargs)

    def execute(self, dataset: OmnesDataArray | None, *args, **kwargs) -> OmnesDataArray | None:
        self._data = read_excel(join(self._path, self._filename), index_col=0, header=0, parse_dates=True,
                                date_format="%d.%m.%Y %H.%M").rename(columns=lambda x: ReadGseDatabase._column_names(x))
        return OmnesDataArray(data=self._data)
